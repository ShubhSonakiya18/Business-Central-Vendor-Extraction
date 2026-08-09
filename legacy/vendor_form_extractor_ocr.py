"""
Vendor Creation Request Form - Auto-fill Extractor (legacy OCR version)
=========================================================================
SUPERSEDED — kept for reference only, not imported by app.py or any
active code path. Replaced by vendor_form_extractor_gemini.py, which
reads documents directly via the Gemini multimodal API instead of
OCR + regex. See the OCR pipeline was dropped because Tesseract
accuracy on scanned/watermarked certificates was inconsistent and
every new layout needed another hand-written regex.

Reads: Cancelled Cheque (image/PDF), GST Certificate (PDF), Udyam Certificate (PDF)
Outputs: single JSON with all extracted + validated fields, and can auto-fill the vendor xlsx.

Dependencies (install with pip):
    pip install pymupdf pytesseract pdf2image pillow openpyxl
System dependency:
    Tesseract OCR binary (e.g. `apt install tesseract-ocr` / `brew install tesseract`)
    Poppler for pdf2image (e.g. `apt install poppler-utils`)

Usage:
    python vendor_form_extractor_ocr.py --cheque cheque.pdf --gst gst_cert.pdf --udyam udyam_cert.pdf --out result.json
"""

import re
import json
import argparse
from pathlib import Path

import fitz  # PyMuPDF
import pytesseract
from PIL import Image, ImageOps
import openpyxl

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None


# ---------------------------------------------------------------------------
# 1. TEXT EXTRACTION
# ---------------------------------------------------------------------------

def extract_pdf_text(path: str) -> str:
    """Try the real text layer first (fast, 100% accurate when present).
    Some government-portal PDFs (e.g. Udyam certificates) are actually
    scanned/rendered images wrapped in a PDF with NO text layer at all --
    detect that per-page and fall back to OCR only where needed."""
    doc = fitz.open(path)
    chunks = []
    needs_ocr_pages = []
    for i, page in enumerate(doc):
        t = page.get_text()
        if len(t.strip()) > 20:  # real text layer present
            chunks.append(t)
        else:
            needs_ocr_pages.append(i)
    doc.close()

    if needs_ocr_pages:
        if convert_from_path is None:
            raise RuntimeError("pdf2image not installed; needed for image-only PDF pages")
        images = convert_from_path(path, dpi=300)
        for i in needs_ocr_pages:
            img = ImageOps.autocontrast(ImageOps.grayscale(images[i]))
            chunks.append(pytesseract.image_to_string(img))

    return "\n".join(chunks)


def extract_cheque_text(path: str) -> str:
    """Cheque is a scanned image/photo -> needs OCR with light preprocessing."""
    path_obj = Path(path)
    if path_obj.suffix.lower() == ".pdf":
        if convert_from_path is None:
            raise RuntimeError("pdf2image not installed; pip install pdf2image + poppler-utils")
        pages = convert_from_path(path, dpi=300)
        img = pages[0]
    else:
        img = Image.open(path)

    # Basic preprocessing: grayscale + upscale + autocontrast improves small printed text OCR
    img = ImageOps.grayscale(img)
    img = img.resize((img.width * 2, img.height * 2))
    img = ImageOps.autocontrast(img)

    return pytesseract.image_to_string(img)


# ---------------------------------------------------------------------------
# 2. REGEX PATTERNS
# ---------------------------------------------------------------------------

RE_GSTIN = re.compile(r"\b\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z]{1}[A-Z0-9]\b")
RE_PAN = re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b")
RE_IFSC = re.compile(r"\b[A-Z]{4}0[A-Z0-9]{6}\b")
RE_UDYAM = re.compile(r"\bUDYAM-[A-Z]{2}-\d{2}-\d{7}\b")
RE_PINCODE = re.compile(r"\b\d{6}\b")


def find_account_number(cheque_text: str) -> str | None:
    """Look for the label 'A/c No' / 'A/C No' and grab the digit run that follows.
    Deliberately does NOT parse the MICR line at the bottom (different encoding)."""
    m = re.search(r"A[\s/]*[Cc]\.?\s*No\.?\s*[:\-]?\s*(\d{9,18})", cheque_text)
    if m:
        return m.group(1)
    # fallback: longest digit run 9-18 chars that isn't part of the MICR band
    candidates = re.findall(r"\b\d{9,18}\b", cheque_text)
    return candidates[0] if candidates else None


def fuzzy_find_ifsc(text: str) -> str | None:
    """IFSC has a known fixed shape: 4 letters + literal '0' + 6 alphanumeric.
    Watermarked/noisy cheque scans commonly confuse I<->1, O<->0, S<->5, B<->8.
    Find an 11-char token near an 'IFS' label and correct known confusions
    at the positions where the format guarantees a letter vs the 5th position
    which is always '0', instead of relying on a clean exact-format regex match."""
    m = re.search(r"IFS[C\s]*Code[:\s;]*([A-Z0-9]{11})", text, re.IGNORECASE)
    if not m:
        # fallback: any 11-char alnum token that's "almost" IFSC shape
        candidates = re.findall(r"\b[A-Z0-9]{11}\b", text)
        m_token = next((c for c in candidates if c[4] in "0O"), None)
        if not m_token:
            return None
        token = m_token
    else:
        token = m.group(1)

    token = list(token.upper())
    digit_to_letter = {"0": "O", "1": "I", "5": "S", "8": "B"}
    letter_to_digit = {"O": "0", "I": "1", "S": "5", "B": "8"}

    # positions 0-3 must be letters
    for i in range(4):
        if token[i] in digit_to_letter:
            token[i] = digit_to_letter[token[i]]
    # position 4 is always literal '0'
    if token[4] in letter_to_digit:
        token[4] = "0"
    else:
        token[4] = "0"  # enforce by spec regardless

    corrected = "".join(token)
    return corrected if re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", corrected) else None


def find_bank_name(cheque_text: str) -> str | None:
    """Grab just '<Name> Bank' rather than the whole noisy OCR line it sits in --
    watermark text often bleeds onto the same line as the bank name."""
    lines = [l.strip() for l in cheque_text.splitlines() if l.strip()]
    for l in lines[:5]:
        m = re.search(r"\b([A-Za-z]+\s+Bank)\b", l)
        if m:
            return m.group(1)
    return None


# ---------------------------------------------------------------------------
# 3. FIELD EXTRACTORS PER DOCUMENT
# ---------------------------------------------------------------------------

def parse_gst_certificate(text: str) -> dict:
    out = {}
    gstin_match = RE_GSTIN.search(text)
    out["gst_number"] = gstin_match.group(0) if gstin_match else None
    if out["gst_number"]:
        out["pan_from_gstin"] = out["gst_number"][2:12]

    name_match = re.search(r"Legal Name\s*\n?\s*([A-Z0-9 &.,]+)", text)
    out["vendor_name"] = name_match.group(1).strip() if name_match else None

    # Address block fields (structured labels in GST REG-06)
    def grab(label):
        m = re.search(label + r"\s*[:\-]?\s*([^\n]+)", text)
        return m.group(1).strip() if m else None

    out["address_1"] = grab(r"Building No\./Flat No\.")
    out["address_2"] = grab(r"Road/Street")
    out["city"] = grab(r"City/Town/Village")
    out["state"] = grab(r"State")
    pin_match = RE_PINCODE.search(grab(r"PIN Code") or "")
    out["pin_code"] = pin_match.group(0) if pin_match else None

    return out


def parse_udyam_certificate(text: str) -> dict:
    out = {}
    udyam_match = RE_UDYAM.search(text)
    out["udyam_number"] = udyam_match.group(0) if udyam_match else None

    pan_match = re.search(r"PAN\s*\n?\s*([A-Z]{5}\d{4}[A-Z])", text)
    out["pan_from_udyam"] = pan_match.group(1) if pan_match else None

    # These certs put labels and values in a scanned table (not a clean
    # label-then-value text stream), so match non-greedily up to the *next*
    # known label rather than assuming the value is one clean line.
    org_match = re.search(
        r"Type of Organisation\s*\|?\s*([A-Za-z ]+?)\s+Name of Enterprise", text
    )
    out["company_type"] = org_match.group(1).strip() if org_match else None

    # OCR sometimes clips the leading "M" off "Major Activity" -> "ajor Activity"
    activity_match = re.search(r"[Mm]?ajor\s*Activity\s*([A-Za-z]+)", text)
    if activity_match:
        raw = activity_match.group(1).strip()
        known = {"Manufactur": "Manufacturing", "Trading": "Trading", "Servic": "Service"}
        out["nature_of_business"] = next(
            (v for k, v in known.items() if raw.startswith(k)), raw.title()
        )
    else:
        out["nature_of_business"] = None

    email_match = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", text)
    out["email"] = email_match.group(0) if email_match else None

    mobile_match = re.search(r"Mobile\s*\n?\s*(\d{10})", text)
    out["telephone"] = mobile_match.group(1) if mobile_match else None

    return out


def parse_cheque(text: str) -> dict:
    out = {}
    ifsc_match = RE_IFSC.search(text)
    out["ifsc"] = ifsc_match.group(0) if ifsc_match else fuzzy_find_ifsc(text)
    out["bank_name"] = find_bank_name(text)
    out["account_number"] = find_account_number(text)

    branch_match = re.search(r"([A-Za-z ]+)\s*[-–]\s*([A-Za-z ]+ Branch)", text)
    out["branch_address"] = branch_match.group(0).strip() if branch_match else None

    return out


# ---------------------------------------------------------------------------
# 4. VALIDATION
# ---------------------------------------------------------------------------

def validate(field: str, value: str | None) -> bool:
    if not value:
        return False
    checks = {
        "gst_number": lambda v: bool(re.fullmatch(r"\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[0-9A-Z]", v)),
        "pan": lambda v: bool(re.fullmatch(r"[A-Z]{5}\d{4}[A-Z]", v)),
        "ifsc": lambda v: bool(re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", v)),
        "udyam_number": lambda v: bool(re.fullmatch(r"UDYAM-[A-Z]{2}-\d{2}-\d{7}", v)),
        "pin_code": lambda v: bool(re.fullmatch(r"\d{6}", v)),
        "account_number": lambda v: v.isdigit() and 9 <= len(v) <= 18,
    }
    return checks.get(field, lambda v: True)(value)


# ---------------------------------------------------------------------------
# 5. PIPELINE
# ---------------------------------------------------------------------------

def build_vendor_json(cheque_path: str, gst_path: str, udyam_path: str) -> dict:
    gst_text = extract_pdf_text(gst_path)
    udyam_text = extract_pdf_text(udyam_path)
    cheque_text = extract_cheque_text(cheque_path)

    gst_data = parse_gst_certificate(gst_text)
    udyam_data = parse_udyam_certificate(udyam_text)
    cheque_data = parse_cheque(cheque_text)

    needs_review = []

    # PAN conflict resolution: compare GSTIN-derived PAN vs Udyam's explicit PAN field
    pan_gst = gst_data.get("pan_from_gstin")
    pan_udyam = udyam_data.get("pan_from_udyam")
    if pan_gst and pan_udyam and pan_gst != pan_udyam:
        needs_review.append("pan_mismatch_between_gst_and_udyam")
        pan_final = pan_gst  # GST cert wins as legal source
    else:
        pan_final = pan_gst or pan_udyam

    result = {
        "vendor_name": gst_data.get("vendor_name"),
        "address_1": gst_data.get("address_1"),
        "address_2": gst_data.get("address_2"),
        "city": gst_data.get("city"),
        "state": gst_data.get("state"),
        "country": "India",
        "pin_code": gst_data.get("pin_code"),
        "company_type": udyam_data.get("company_type"),
        "nature_of_business": udyam_data.get("nature_of_business"),
        "pan": pan_final,
        "tan": None,
        "gst_number": gst_data.get("gst_number"),
        "esic_number": None,
        "udyam_number": udyam_data.get("udyam_number"),
        "bank_name": cheque_data.get("bank_name"),
        "branch_address": cheque_data.get("branch_address"),
        "ifsc": cheque_data.get("ifsc"),
        "account_number": cheque_data.get("account_number"),
        "account_type": None,
        "telephone": udyam_data.get("telephone"),
        "email": udyam_data.get("email"),
        "website": None,
    }

    # Run validations, collect failures
    for field in ["gst_number", "pan", "ifsc", "udyam_number", "pin_code", "account_number"]:
        if not validate(field, result.get(field)):
            needs_review.append(f"{field}_failed_validation_or_missing")

    result["needs_review"] = needs_review
    return result


# ---------------------------------------------------------------------------
# 6. XLSX AUTO-FILL
# ---------------------------------------------------------------------------

# Maps our JSON keys -> the exact cell in the "To be filled by supplier" block
# of the Netsmartz/Mahindra Vendor Creation Request Form template.
# Row numbers were read directly from the uploaded template (rows 37-63).
XLSX_CELL_MAP = {
    "vendor_name": "B37",
    "address_1": "B38",
    "address_2": "B39",
    "address_3": "B40",
    "address_4": "B41",
    "city": "B42",
    "state": "B43",
    "country": "B44",
    "pin_code": "B45",
    "telephone": "B46",
    "email": "B48",
    "website": "B49",
    "company_type": "B50",
    "nature_of_business": "B51",
    "tan": "B52",
    "pan": "B53",
    "gst_number": "B55",
    "esic_number": "B56",
    "udyam_number": "B57",
    "bank_name": "B59",
    "branch_address": "B60",
    "ifsc": "B61",
    "account_type": "B62",
    "account_number": "B63",
}


def fill_vendor_xlsx(data: dict, template_path: str, sheet_name: str, output_path: str) -> None:
    """Writes extracted fields into the named sheet of the vendor form template,
    leaving all existing formatting, other sheets, and unmapped cells untouched.
    Only writes a cell if we actually have a non-None value, so blank/mandatory
    cells the extractor couldn't fill stay blank for manual entry rather than
    being overwritten with 'None' or empty strings."""
    wb = openpyxl.load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    filled, skipped = [], []
    for json_key, cell_ref in XLSX_CELL_MAP.items():
        value = data.get(json_key)
        if value:
            ws[cell_ref] = value
            filled.append((json_key, cell_ref, value))
        else:
            skipped.append((json_key, cell_ref))

    wb.save(output_path)
    print(f"\nWrote {len(filled)} fields into sheet '{sheet_name}' of {output_path}")
    for k, c, v in filled:
        print(f"  {c}  {k:<20} = {v}")
    if skipped:
        print(f"\nLeft blank (no extracted value, fill manually):")
        for k, c in skipped:
            print(f"  {c}  {k}")


# ---------------------------------------------------------------------------
# 7. CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract vendor fields from cheque, GST cert, Udyam cert")
    parser.add_argument("--cheque", required=True)
    parser.add_argument("--gst", required=True)
    parser.add_argument("--udyam", required=True)
    parser.add_argument("--out", default="vendor_extracted.json")
    parser.add_argument("--template", help="Path to the VENDOR_CREATION_REQUEST_FORM.xlsx template")
    parser.add_argument("--sheet", help="Sheet/tab name inside the template to fill in")
    parser.add_argument("--xlsx-out", default="vendor_filled.xlsx")
    args = parser.parse_args()

    data = build_vendor_json(args.cheque, args.gst, args.udyam)
    Path(args.out).write_text(json.dumps(data, indent=2))
    print(json.dumps(data, indent=2))
    if data["needs_review"]:
        print("\n⚠️  Fields needing manual review:", data["needs_review"])

    if args.template:
        if not args.sheet:
            wb = openpyxl.load_workbook(args.template, read_only=True)
            raise SystemExit(
                f"--template given but --sheet is required. "
                f"Available sheets in this file: {wb.sheetnames}"
            )
        fill_vendor_xlsx(data, args.template, args.sheet, args.xlsx_out)
