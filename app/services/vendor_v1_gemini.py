"""
Vendor Creation Request Form - Auto-fill Extractor (V1, Gemini)
=================================================================
Reads: Cancelled Cheque (image/PDF), GST Certificate (PDF), Udyam Certificate (PDF)
Uses the Gemini API (multimodal) to read each document directly.

Dependencies:
    pip install google-genai openpyxl

Environment:
    GEMINI_API_KEY must be set (see app/core/config.py, .env.example).

Usage:
    python -m app.services.vendor_v1_gemini --cheque cheque.pdf --gst gst_cert.pdf --udyam udyam_cert.pdf --out result.json
"""

import os
import re
import json
import argparse
import mimetypes
from pathlib import Path

from google import genai
from google.genai import types
from google.genai.errors import ServerError
import openpyxl
from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential


MODEL = "gemini-flash-lite-latest"

# Gemini occasionally returns 503 ("high demand") or 429 (rate limit) under
# load; these are transient and worth a few retries with backoff. Anything
# else (bad request, invalid key, etc.) should fail immediately.
def _is_transient(exc: Exception) -> bool:
    return isinstance(exc, ServerError) or (
        isinstance(exc, Exception) and getattr(exc, "code", None) == 429
    )


retry_transient = retry(
    retry=retry_if_exception(_is_transient),
    stop=stop_after_attempt(4),
    wait=wait_exponential(multiplier=2, min=2, max=60),
    reraise=True,
)

GST_SCHEMA = {
    "type": "object",
    "properties": {
        "gst_number": {"type": "string"},
        "pan_from_gstin": {"type": "string"},
        "vendor_name": {"type": "string"},
        "address_1": {"type": "string"},
        "address_2": {"type": "string"},
        "city": {"type": "string"},
        "state": {"type": "string"},
        "pin_code": {"type": "string"},
    },
}

UDYAM_SCHEMA = {
    "type": "object",
    "properties": {
        "udyam_number": {"type": "string"},
        "pan_from_udyam": {"type": "string"},
        "company_type": {"type": "string"},
        "nature_of_business": {"type": "string"},
        "email": {"type": "string"},
        "telephone": {"type": "string"},
    },
}

CHEQUE_SCHEMA = {
    "type": "object",
    "properties": {
        "bank_name": {"type": "string"},
        "branch_address": {"type": "string"},
        "ifsc": {"type": "string"},
        "account_number": {"type": "string"},
    },
}


def _upload(client: genai.Client, path: str):
    mime, _ = mimetypes.guess_type(path)
    return client.files.upload(file=path, config={"mime_type": mime or "application/pdf"})


@retry_transient
def _generate(client: genai.Client, file_ref, prompt: str, schema: dict):
    return client.models.generate_content(
        model=MODEL,
        contents=[file_ref, prompt],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=schema,
        ),
    )


def _extract(client: genai.Client, path: str, prompt: str, schema: dict) -> dict:
    file_ref = _upload(client, path)
    response = _generate(client, file_ref, prompt, schema)
    try:
        return json.loads(response.text)
    except (json.JSONDecodeError, TypeError):
        return {}


def parse_gst_certificate(client: genai.Client, path: str) -> dict:
    prompt = (
        "This is an Indian GST Registration Certificate (REG-06). Extract these fields "
        "exactly as printed: gst_number (the 15-char GSTIN), pan_from_gstin (10-char PAN "
        "embedded in the GSTIN, characters 3-12), vendor_name (Legal Name of Business), "
        "address_1 (Building No./Flat No. + Floor/Name of Premises), address_2 (Road/Street/"
        "Locality/Area), city (City/Town/Village), state, pin_code. Use null for any field "
        "not found."
    )
    return _extract(client, path, prompt, GST_SCHEMA)


def parse_udyam_certificate(client: genai.Client, path: str) -> dict:
    prompt = (
        "This is an Indian Udyam Registration Certificate. Extract these fields exactly as "
        "printed: udyam_number (format UDYAM-XX-00-0000000), pan_from_udyam (10-char PAN), "
        "company_type (Type of Organisation), nature_of_business (Major Activity: "
        "Manufacturing/Trading/Service), email, telephone (mobile number). Use null for any "
        "field not found."
    )
    return _extract(client, path, prompt, UDYAM_SCHEMA)


def parse_cheque(client: genai.Client, path: str) -> dict:
    prompt = (
        "This is a scanned/photographed Indian bank cancelled cheque. Extract these fields "
        "exactly as printed: bank_name, branch_address, ifsc (11-char IFSC code, 4 letters + "
        "'0' + 6 alphanumeric), account_number (the printed account number near 'A/c No.', "
        "NOT the MICR line at the bottom). Use null for any field not found."
    )
    return _extract(client, path, prompt, CHEQUE_SCHEMA)


# ---------------------------------------------------------------------------
# VALIDATION
# ---------------------------------------------------------------------------

def validate(field: str, value) -> bool:
    if not value:
        return False
    checks = {
        "gst_number": lambda v: bool(re.fullmatch(r"\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[0-9A-Z]", v)),
        "pan": lambda v: bool(re.fullmatch(r"[A-Z]{5}\d{4}[A-Z]", v)),
        "ifsc": lambda v: bool(re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", v)),
        "udyam_number": lambda v: bool(re.fullmatch(r"UDYAM-[A-Z]{2}-\d{2}-\d{7}", v)),
        "pin_code": lambda v: bool(re.fullmatch(r"\d{6}", v)),
        "account_number": lambda v: v.isdigit() and 9 <= len(v) <= 18,
    }
    return checks.get(field, lambda v: True)(value)


# ---------------------------------------------------------------------------
# PIPELINE
# ---------------------------------------------------------------------------

def build_vendor_json(cheque_path: str, gst_path: str, udyam_path: str) -> dict:
    client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))

    gst_data = parse_gst_certificate(client, gst_path)
    udyam_data = parse_udyam_certificate(client, udyam_path)
    cheque_data = parse_cheque(client, cheque_path)

    needs_review = []

    pan_gst = gst_data.get("pan_from_gstin")
    pan_udyam = udyam_data.get("pan_from_udyam")
    if pan_gst and pan_udyam and pan_gst != pan_udyam:
        needs_review.append("pan_mismatch_between_gst_and_udyam")
        pan_final = pan_gst  # GST cert wins as legal source
    else:
        pan_final = pan_gst or pan_udyam

    result = {
        "vendor_name": gst_data.get("vendor_name"),
        "address_1": gst_data.get("address_1"),
        "address_2": gst_data.get("address_2"),
        "city": gst_data.get("city"),
        "state": gst_data.get("state"),
        "country": "India",
        "pin_code": gst_data.get("pin_code"),
        "company_type": udyam_data.get("company_type"),
        "nature_of_business": udyam_data.get("nature_of_business"),
        "pan": pan_final,
        "tan": None,
        "gst_number": gst_data.get("gst_number"),
        "esic_number": None,
        "udyam_number": udyam_data.get("udyam_number"),
        "bank_name": cheque_data.get("bank_name"),
        "branch_address": cheque_data.get("branch_address"),
        "ifsc": cheque_data.get("ifsc"),
        "account_number": cheque_data.get("account_number"),
        "account_type": None,
        "telephone": udyam_data.get("telephone"),
        "email": udyam_data.get("email"),
        "website": None,
    }

    for field in ["gst_number", "pan", "ifsc", "udyam_number", "pin_code", "account_number"]:
        if not validate(field, result.get(field)):
            needs_review.append(f"{field}_failed_validation_or_missing")

    result["needs_review"] = needs_review
    return result


# ---------------------------------------------------------------------------
# XLSX AUTO-FILL
# ---------------------------------------------------------------------------

XLSX_CELL_MAP = {
    "vendor_name": "B37",
    "address_1": "B38",
    "address_2": "B39",
    "address_3": "B40",
    "address_4": "B41",
    "city": "B42",
    "state": "B43",
    "country": "B44",
    "pin_code": "B45",
    "telephone": "B46",
    "email": "B48",
    "website": "B49",
    "company_type": "B50",
    "nature_of_business": "B51",
    "tan": "B52",
    "pan": "B53",
    "gst_number": "B55",
    "esic_number": "B56",
    "udyam_number": "B57",
    "bank_name": "B59",
    "branch_address": "B60",
    "ifsc": "B61",
    "account_type": "B62",
    "account_number": "B63",
}


def fill_vendor_xlsx(data: dict, template_path: str, sheet_name: str, output_path: str) -> None:
    wb = openpyxl.load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    filled, skipped = [], []
    for json_key, cell_ref in XLSX_CELL_MAP.items():
        value = data.get(json_key)
        ws[cell_ref] = value if value else None
        if not value:
            ws[cell_ref].hyperlink = None
        if value:
            filled.append((json_key, cell_ref, value))
        else:
            skipped.append((json_key, cell_ref))

    wb.save(output_path)
    print(f"\nWrote {len(filled)} fields into sheet '{sheet_name}' of {output_path}")
    for k, c, v in filled:
        print(f"  {c}  {k:<20} = {v}")
    if skipped:
        print(f"\nLeft blank (no extracted value, fill manually):")
        for k, c in skipped:
            print(f"  {c}  {k}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract vendor fields from cheque, GST cert, Udyam cert using Gemini")
    parser.add_argument("--cheque", required=True)
    parser.add_argument("--gst", required=True)
    parser.add_argument("--udyam", required=True)
    parser.add_argument("--out", default="vendor_extracted.json")
    parser.add_argument("--template", help="Path to the VENDOR_CREATION_REQUEST_FORM.xlsx template")
    parser.add_argument("--sheet", help="Sheet/tab name inside the template to fill in")
    parser.add_argument("--xlsx-out", default="vendor_filled.xlsx")
    args = parser.parse_args()

    if not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit("GEMINI_API_KEY environment variable is not set.")

    data = build_vendor_json(args.cheque, args.gst, args.udyam)
    Path(args.out).write_text(json.dumps(data, indent=2))
    print(json.dumps(data, indent=2))
    if data["needs_review"]:
        print("\n⚠️  Fields needing manual review:", data["needs_review"])

    if args.template:
        if not args.sheet:
            wb = openpyxl.load_workbook(args.template, read_only=True)
            raise SystemExit(
                f"--template given but --sheet is required. "
                f"Available sheets in this file: {wb.sheetnames}"
            )
        fill_vendor_xlsx(data, args.template, args.sheet, args.xlsx_out)

        from app.services.vendor_v1_verify import verify_vendor_excel
        verify_vendor_excel(data, args.xlsx_out, args.sheet)
