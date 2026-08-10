"""
V2 Excel Mapper
===============
Writes the canonical vendor JSON into a template using a YAML cell map from
config/excel_mappings/. Supporting a new template means adding a YAML file.

Deliberately knows nothing about OCR, PaddleOCR or the semantic engine -- it
takes a plain dict. That keeps it usable with V1's output too, and makes it
testable without running any of the extraction stack.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
import yaml

from .config_loader import CONFIG_DIR

logger = logging.getLogger(__name__)

MAPPINGS_DIR = CONFIG_DIR / "excel_mappings"


class MappingError(ValueError):
    pass


@dataclass
class CellMapping:
    field: str
    cell: str
    sheet: Optional[str] = None


class ExcelMapper:
    def __init__(self, name: str, mappings: list[CellMapping], description: str = "", blank_value=None):
        self.name = name
        self.mappings = mappings
        self.description = description
        self.blank_value = blank_value

    # -- loading ------------------------------------------------------------

    @classmethod
    def load(cls, name: str, directory: Optional[Path] = None) -> "ExcelMapper":
        directory = directory or MAPPINGS_DIR
        path = directory / f"{name}.yaml" if not name.endswith(".yaml") else Path(name)
        if not path.exists():
            available = sorted(p.stem for p in directory.glob("*.yaml"))
            raise MappingError(f"Excel mapping {name!r} not found. Available: {available}")

        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        raw = data.get("fields") or {}
        if not raw:
            raise MappingError(f"{path.name} defines no fields")

        mappings: list[CellMapping] = []
        seen: dict[tuple[Optional[str], str], str] = {}
        for field, cfg in raw.items():
            if isinstance(cfg, str):
                cfg = {"cell": cfg}
            cell = (cfg or {}).get("cell")
            if not cell:
                raise MappingError(f"{path.name}: field {field!r} has no cell")
            sheet = (cfg or {}).get("sheet")

            # Two fields writing the same cell means one silently overwrites
            # the other -- always a config mistake, and invisible at runtime.
            key = (sheet, str(cell).upper())
            if key in seen:
                raise MappingError(
                    f"{path.name}: {field!r} and {seen[key]!r} both map to cell {cell}"
                )
            seen[key] = field

            mappings.append(CellMapping(field=field, cell=str(cell).upper(), sheet=sheet))

        return cls(
            name=data.get("template", path.stem),
            mappings=mappings,
            description=data.get("description", ""),
            blank_value=data.get("blank_value"),
        )

    @staticmethod
    def available(directory: Optional[Path] = None) -> list[str]:
        directory = directory or MAPPINGS_DIR
        return sorted(p.stem for p in directory.glob("*.yaml")) if directory.exists() else []

    # -- writing ------------------------------------------------------------

    def fill(
        self,
        data: dict,
        template_path: str,
        output_path: str,
        sheet_name: Optional[str] = None,
    ) -> int:
        """Write mapped values into the template and save to output_path.

        Returns the number of non-empty values written. Everything the template
        already contains -- formatting, other sheets, unmapped cells -- is left
        untouched, because the file goes on to a human reviewer who expects the
        form they recognise.
        """
        workbook = openpyxl.load_workbook(template_path)
        written = 0

        for mapping in self.mappings:
            target_sheet = mapping.sheet or sheet_name
            if target_sheet is None:
                target_sheet = workbook.sheetnames[0]
            if target_sheet not in workbook.sheetnames:
                raise MappingError(
                    f"Sheet {target_sheet!r} not found. Available: {workbook.sheetnames}"
                )

            worksheet = workbook[target_sheet]
            value = data.get(mapping.field)
            cell = worksheet[mapping.cell]

            if value:
                cell.value = value
                written += 1
            else:
                cell.value = self.blank_value
                # A blanked template cell can still carry a hyperlink from the
                # original form, which renders as clickable empty text.
                cell.hyperlink = None

        workbook.save(output_path)
        logger.info("Wrote %d field(s) into %s", written, output_path)
        return written

    def cells_for(self, sheet_name: Optional[str] = None) -> list[tuple[str, str, Optional[str]]]:
        """(field, cell, sheet) triples, for the verifier."""
        return [(m.field, m.cell, m.sheet or sheet_name) for m in self.mappings]
