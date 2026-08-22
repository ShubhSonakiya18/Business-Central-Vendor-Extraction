"""
V2 Excel Verification
=====================
Preserves V1's read-back check, driven by the YAML cell map instead of the
hardcoded one: save the workbook, reopen it, read every mapped cell, compare
against the canonical JSON it came from, colour PASS green and FAIL red, and
write verification_report.json.

Reopening rather than trusting the in-memory workbook is the point. It catches
a cell that silently failed to write, a value coerced to a different type by
Excel, or a mapping pointing at a merged cell whose value lands elsewhere.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

from .excel_mapper import ExcelMapper

logger = logging.getLogger(__name__)

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _normalize(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # openpyxl hands back numeric cells as floats, so a PIN code written as
    # "700019" can read back as "700019.0" and fail a comparison it should pass.
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.lower()


def verify_excel(
    data: dict,
    xlsx_path: str,
    mapper: ExcelMapper,
    sheet_name: Optional[str] = None,
    report_path: Optional[str] = None,
    highlight: bool = True,
) -> list[dict]:
    """Compare every mapped cell in the saved workbook against `data`."""
    workbook = openpyxl.load_workbook(xlsx_path)
    report: list[dict] = []

    for field, cell_ref, sheet in mapper.cells_for(sheet_name):
        target = sheet or workbook.sheetnames[0]
        if target not in workbook.sheetnames:
            raise ValueError(f"Sheet {target!r} not found. Available: {workbook.sheetnames}")
        worksheet = workbook[target]

        expected = data.get(field)
        actual = worksheet[cell_ref].value
        status = "PASS" if _normalize(expected) == _normalize(actual) else "FAIL"

        report.append(
            {
                "field": field,
                "sheet": target,
                "cell": cell_ref,
                "expected": expected,
                "actual": actual,
                "status": status,
            }
        )

        if highlight:
            worksheet[cell_ref].fill = PASS_FILL if status == "PASS" else FAIL_FILL

    if highlight:
        workbook.save(xlsx_path)

    if report_path:
        p = Path(report_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    return report


def summarize(report: list[dict]) -> dict:
    total = len(report)
    passed = sum(1 for e in report if e["status"] == "PASS")
    return {
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "success_rate": round(passed / total * 100) if total else 0,
    }


def print_report(report: list[dict]) -> None:
    print("=" * 60)
    print("Vendor Form Verification")
    print("=" * 60)
    for entry in report:
        icon = "PASS" if entry["status"] == "PASS" else "FAIL"
        dots = "." * max(1, 26 - len(entry["field"]))
        print(f"  [{icon}] {entry['field']} {dots} {entry['cell']}")
        if entry["status"] == "FAIL":
            print(f"         expected: {entry['expected']!r}")
            print(f"         actual:   {entry['actual']!r}")

    s = summarize(report)
    print("-" * 60)
    print(f"  checked {s['total']}, passed {s['passed']}, failed {s['failed']} "
          f"({s['success_rate']}%)")
    print("=" * 60)
