"""
FastAPI wrapper around the vendor extraction / Excel-fill / verification
pipeline.

Fully local: PaddleOCR plus a YAML-driven semantic engine. No network calls,
no API key.

Run with:
    uvicorn app:app --reload
"""

import json
import logging
import re
import shutil
import uuid
from pathlib import Path

from fastapi import FastAPI, Request, UploadFile, File, Form
from typing import List, Optional
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(name)s: %(message)s",
)
logger = logging.getLogger("vendor_extractor")

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Vendor Form Extractor")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")


def _save_upload(upload: UploadFile, dest_dir: Path) -> Path:
    dest = dest_dir / upload.filename
    with dest.open("wb") as f:
        shutil.copyfileobj(upload.file, f)
    return dest


def _run_state_path(run_id: str) -> Path:
    return OUTPUT_DIR / run_id / "run_state.json"


def _save_run(run_id: str, state: dict) -> None:
    """Persist a completed run so results survive a restart.

    An in-memory dict would grow without bound and, under more than one uvicorn
    worker, would strand a user on a worker that never saw their run. Everything
    else about the run is already written under outputs/<run_id>/, so the state
    belongs there too.
    """
    path = _run_state_path(run_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def _load_run(run_id: str) -> Optional[dict]:
    """Return a persisted run, or None if the id is unknown.

    run_id comes straight from the URL, so it is confined to the hex ids this
    app generates before it is used as a path segment -- otherwise a crafted id
    could walk out of the outputs directory.
    """
    if not re.fullmatch(r"[0-9a-f]{10}", run_id):
        return None
    path = _run_state_path(run_id)
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        logger.exception("Could not read run state for %s", run_id)
        return None


@app.get("/")
def v2_home(request: Request):
    from vendor_extractor.excel.excel_mapper import ExcelMapper

    return templates.TemplateResponse(
        request, "index.html", {"mappings": ExcelMapper.available() or ["vendor_creation_v1"]}
    )


def _v2_error(request: Request, message: str, detail: str = "", hint: str = "", sheets=None):
    return templates.TemplateResponse(
        request,
        "error.html",
        {"message": message, "detail": detail, "hint": hint, "available_sheets": sheets or []},
        status_code=400,
    )


@app.post("/v2/process-vendor")
def v2_process(
    request: Request,
    documents: List[UploadFile] = File(...),
    vendor_template: Optional[UploadFile] = File(None),
    sheet_names: List[str] = Form(default=[]),
    mapping: str = Form("vendor_creation_v1"),
    models: str = Form("small"),
):
    """Run the local pipeline over an arbitrary set of vendor documents.

    Everything happens in-process: no queue, no external service. Uploads and
    outputs are kept per run so a result can be re-downloaded and audited.
    """
    import time
    import traceback

    import openpyxl

    from vendor_extractor.ingest.document_loader import load_documents
    from vendor_extractor.excel.excel_mapper import ExcelMapper
    from vendor_extractor.ingest.ocr_engine import OCREngine
    from vendor_extractor.pipeline import extract_from_document_set
    from vendor_extractor.excel.verifier import summarize, verify_excel

    run_id = uuid.uuid4().hex[:10]
    run_upload_dir = UPLOAD_DIR / run_id
    run_output_dir = OUTPUT_DIR / run_id
    run_upload_dir.mkdir(parents=True, exist_ok=True)
    run_output_dir.mkdir(parents=True, exist_ok=True)

    saved = [_save_upload(doc, run_upload_dir) for doc in documents if doc.filename]
    template_path = (
        _save_upload(vendor_template, run_upload_dir)
        if vendor_template is not None and vendor_template.filename
        else None
    )

    if not saved:
        return _v2_error(
            request,
            "No documents were uploaded.",
            hint="Select at least one PDF, DOCX or image file.",
        )

    # Validate the requested sheets against the uploaded workbook before doing
    # two minutes of OCR. The sheet checkboxes on the form are a fixed list, so
    # a template whose tabs are named differently would otherwise fail only at
    # the very end, after all the expensive work was already done.
    available_sheets: list[str] = []
    if template_path is not None:
        try:
            probe = openpyxl.load_workbook(template_path, read_only=True)
            available_sheets = list(probe.sheetnames)
            probe.close()
        except Exception as exc:
            return _v2_error(
                request,
                "That Excel template could not be opened.",
                detail=f"{type(exc).__name__}: {exc}",
                hint="The file must be a valid .xlsx workbook.",
            )

        missing = [s for s in sheet_names if s not in available_sheets]
        if missing:
            return _v2_error(
                request,
                f"The template does not contain the sheet(s) you selected: "
                f"{', '.join(missing)}.",
                hint="Tick only sheets that exist in your template, or tick none "
                     "to fill the first sheet.",
                sheets=available_sheets,
            )

    try:
        t0 = time.perf_counter()
        engine = OCREngine(
            det_model=f"PP-OCRv6_{models}_det",
            rec_model=f"PP-OCRv6_{models}_rec",
        )
        doc_set = load_documents(saved, engine=engine)
        load_seconds = time.perf_counter() - t0
    except Exception:
        logger.exception("run %s: document loading failed", run_id)
        return _v2_error(
            request,
            "Failed while reading the uploaded documents.",
            detail=traceback.format_exc(limit=6),
            hint="Check the terminal running uvicorn for the full traceback.",
        )

    if not doc_set.documents:
        return _v2_error(
            request,
            "None of the uploaded files could be read.",
            hint="Supported types are PDF, DOCX, PNG, JPG, TIFF, BMP and WEBP.",
        )

    doc_set.save_json(run_output_dir / "document_set.json")

    result = extract_from_document_set(doc_set)
    canonical = result.canonical()

    result.save_json(run_output_dir / "extraction.json")
    json_path = run_output_dir / "result.json"
    json_path.write_text(json.dumps(canonical, indent=2, ensure_ascii=False), encoding="utf-8")

    files = {
        "json": str(json_path),
        "extraction": str(run_output_dir / "extraction.json"),
        "spans": str(run_output_dir / "document_set.json"),
    }
    report: list[dict] = []
    verification = None

    if template_path is not None:
        try:
            mapper = ExcelMapper.load(mapping)
            xlsx_path = run_output_dir / "vendor_filled.xlsx"

            # Each selected tab is filled in turn, re-reading the previous
            # output so every sheet ends up in one workbook.
            targets = sheet_names or [None]
            source = template_path
            for sheet in targets:
                mapper.fill(canonical, str(source), str(xlsx_path), sheet_name=sheet)
                source = xlsx_path

            for sheet in targets:
                sheet_report = verify_excel(
                    canonical, str(xlsx_path), mapper, sheet_name=sheet,
                    report_path=str(run_output_dir / f"verification_{sheet or 'default'}.json"),
                )
                report.extend(sheet_report)

            report_path = run_output_dir / "verification_report.json"
            report_path.write_text(
                json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
            )

            verification = summarize(report)
            files["xlsx"] = str(xlsx_path)
            files["report"] = str(report_path)
        except Exception as exc:
            # Extraction already succeeded and cost real time; losing it to an
            # Excel problem would be the wrong trade. Report the failure and
            # keep the JSON results available.
            logger.exception("run %s: Excel fill/verify failed", run_id)
            return _v2_error(
                request,
                f"Documents were extracted successfully, but filling the Excel "
                f"template failed: {exc}",
                detail=traceback.format_exc(limit=6),
                hint=f"The extracted data was still saved to outputs/{run_id}/result.json",
                sheets=available_sheets,
            )

    _save_run(run_id, {
        "fields": {k: v.to_dict() for k, v in result.fields.items()},
        "needs_review": result.needs_review,
        "documents": result.documents,
        "summary": result.summary(),
        "report": report,
        "verification": verification,
        "files": files,
        "timings": {
            "load": round(load_seconds, 1),
            "extract": round(result.duration_s, 2),
            "total": round(time.perf_counter() - t0, 1),
        },
    })

    logger.info(
        "run %s complete: %d/%d fields filled in %.1fs",
        run_id, result.summary().get("filled", 0), len(result.fields),
        time.perf_counter() - t0,
    )
    return RedirectResponse(url=f"/v2/results/{run_id}", status_code=303)


@app.get("/v2/results/{run_id}")
def v2_results(request: Request, run_id: str):
    run = _load_run(run_id)
    if not run:
        return RedirectResponse(url="/")
    return templates.TemplateResponse(
        request,
        "results.html",
        {
            "run_id": run_id,
            "fields": run["fields"],
            "needs_review": run["needs_review"],
            "documents": run["documents"],
            "summary": run["summary"],
            "report": run["report"],
            "verification": run["verification"],
            "downloads": run["files"],
            "timings": run["timings"],
        },
    )


@app.get("/v2/download/{run_id}/{kind}")
def v2_download(run_id: str, kind: str):
    run = _load_run(run_id)
    if not run or kind not in run["files"]:
        return RedirectResponse(url="/")
    path = run["files"][kind]
    return FileResponse(path, filename=Path(path).name)


@app.post("/v2/template-sheets")
async def v2_template_sheets(vendor_template: UploadFile = File(...)):
    """Return the sheet names inside an uploaded workbook.

    Lets the upload form offer the tabs that actually exist in the user's
    template instead of a hardcoded list. Reads from memory; nothing is kept.
    """
    import io

    import openpyxl

    try:
        content = await vendor_template.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheets = list(workbook.sheetnames)
        workbook.close()
        return {"sheets": sheets}
    except Exception as exc:
        return {"sheets": [], "error": f"{type(exc).__name__}: {exc}"}


@app.get("/v2/health")
def v2_health():
    """Confirms the local stack is importable and reports what is configured."""
    from vendor_extractor.config_loader import load_config
    from vendor_extractor.excel.excel_mapper import ExcelMapper

    dictionary, rules = load_config()
    return {
        "status": "ok",
        "mode": "local",
        "fields": len(dictionary),
        "validators": len(rules.validators),
        "excel_mappings": ExcelMapper.available(),
    }
