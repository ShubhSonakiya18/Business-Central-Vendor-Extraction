"""
Vendor Excel Verification
==========================
Post-processing step: reopens a filled vendor Excel template and checks each
mapped cell against the extracted JSON it was supposed to be filled from.

Does not touch extraction or Excel-writing logic. Reuses XLSX_CELL_MAP from
vendor_form_extractor_gemini.py so the two stay in sync.

Usage (standalone):
    python verify_vendor_excel.py --json result.json --xlsx vendor_filled.xlsx --sheet Sheet1

Usage (as a module):
    from verify_vendor_excel import verify_vendor_excel
    report = verify_vendor_excel(data, "vendor_filled.xlsx", "Sheet1")
"""

import json
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from vendor_form_extractor_gemini import XLSX_CELL_MAP

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FIELD_LABELS = {
    "vendor_name": "Vendor Name",
    "address_1": "Address 1",
    "address_2": "Address 2",
    "address_3": "Address 3",
    "address_4": "Address 4",
    "city": "City",
    "state": "State",
    "country": "Country",
    "pin_code": "Pin Code",
    "telephone": "Telephone",
    "email": "Email",
    "website": "Website",
    "company_type": "Company Type",
    "nature_of_business": "Nature of Business",
    "tan": "TAN",
    "pan": "PAN",
    "gst_number": "GST Number",
    "esic_number": "ESIC Number",
    "udyam_number": "Udyam Number",
    "bank_name": "Bank Name",
    "branch_address": "Branch Address",
    "ifsc": "IFSC",
    "account_type": "Account Type",
    "account_number": "Account Number",
}


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def verify_vendor_excel(
    data: dict,
    xlsx_path: str,
    sheet_name: str,
    report_path: str = "verification_report.json",
    highlight: bool = True,
) -> list[dict]:
    """Reopens xlsx_path, compares each mapped cell against data, writes
    verification_report.json, and (optionally) highlights cells in place."""
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    report = []
    for json_key, cell_ref in XLSX_CELL_MAP.items():
        expected = data.get(json_key)
        actual = ws[cell_ref].value
        status = "PASS" if _normalize(expected) == _normalize(actual) else "FAIL"

        report.append({
            "field": json_key,
            "cell": cell_ref,
            "expected": expected,
            "actual": actual,
            "status": status,
        })

        if highlight:
            ws[cell_ref].fill = PASS_FILL if status == "PASS" else FAIL_FILL

    if highlight:
        wb.save(xlsx_path)

    Path(report_path).write_text(json.dumps(report, indent=2))
    _print_summary(report)
    return report


def _print_summary(report: list[dict]) -> None:
    print("=" * 50)
    print("Vendor Form Verification")
    print("=" * 50)
    print()

    for entry in report:
        label = FIELD_LABELS.get(entry["field"], entry["field"])
        icon = "✅" if entry["status"] == "PASS" else "❌"
        dots = "." * max(1, 30 - len(label))
        print(f"{icon} {label} {dots} {entry['status']}")
        if entry["status"] == "FAIL":
            print(f"   Expected: {entry['expected']}")
            print(f"   Actual:   {entry['actual']}")

    total = len(report)
    passed = sum(1 for e in report if e["status"] == "PASS")
    failed = total - passed
    rate = (passed / total * 100) if total else 0

    print()
    print("-" * 50)
    print(f"Fields Checked : {total}")
    print(f"Passed         : {passed}")
    print(f"Failed         : {failed}")
    print(f"Success Rate   : {rate:.0f}%")
    print("=" * 50)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify a filled vendor Excel against its extracted JSON")
    parser.add_argument("--json", required=True, help="Path to extracted vendor JSON")
    parser.add_argument("--xlsx", required=True, help="Path to filled vendor xlsx")
    parser.add_argument("--sheet", required=True, help="Sheet name inside the xlsx")
    parser.add_argument("--report", default="verification_report.json")
    parser.add_argument("--no-highlight", action="store_true", help="Skip coloring cells in the xlsx")
    args = parser.parse_args()

    data = json.loads(Path(args.json).read_text())
    verify_vendor_excel(data, args.xlsx, args.sheet, args.report, highlight=not args.no_highlight)
