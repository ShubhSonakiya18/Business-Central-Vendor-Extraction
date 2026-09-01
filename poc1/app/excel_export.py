"""Generate the two-column vendor-onboarding Excel sheet.

Left column = field labels (bold), right column = values, one row per
field, in the exact order from EXCEL_FIELD_ORDER. "Customer Agreement..."
and "Type" are always left blank for manual entry unless a human has
already filled them in via PUT /extraction.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.crud import list_fields
from app.enums import EXCEL_FIELD_LABELS, EXCEL_FIELD_ORDER, MANUAL_ONLY_FIELDS
from app.models import Vendor

_THIN = Side(style="thin", color="B7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LABEL_FONT = Font(bold=True)
_HEADER_FONT = Font(bold=True, size=13)


def build_vendor_workbook(db: Session, vendor: Vendor) -> Workbook:
    field_rows = {f.field_name: f for f in list_fields(db, vendor.id)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Onboarding"

    ws["A1"] = "Vendor Onboarding Sheet"
    ws["A1"].font = _HEADER_FONT
    ws.merge_cells("A1:B1")

    row = 3
    for field_name in EXCEL_FIELD_ORDER:
        label = EXCEL_FIELD_LABELS[field_name]
        record = field_rows.get(field_name.value)
        if field_name in MANUAL_ONLY_FIELDS:
            value = record.value if (record and record.is_human_edited) else ""
        else:
            value = record.value if record else ""

        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value or "")

        label_cell.font = _LABEL_FONT
        label_cell.border = _BORDER
        label_cell.alignment = Alignment(vertical="center", wrap_text=True)
        value_cell.border = _BORDER
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    ws.column_dimensions[get_column_letter(1)].width = 42
    ws.column_dimensions[get_column_letter(2)].width = 55
    ws.freeze_panes = "A3"

    return wb


def export_vendor_excel(db: Session, vendor: Vendor, out_path: Path) -> Path:
    wb = build_vendor_workbook(db, vendor)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
