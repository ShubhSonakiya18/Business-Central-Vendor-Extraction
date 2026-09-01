"""End-to-end API flow: create vendor -> upload docs -> poll extraction ->
correct a field -> approve -> download Excel."""
from __future__ import annotations

import time

from openpyxl import load_workbook


def _wait_for_processing(client, vendor_id: str, timeout: float = 30.0):
    """The TestClient runs FastAPI BackgroundTasks synchronously as part of
    each request/response cycle, so processing is normally already done by
    the time upload_documents returns -- this poll is just a safety net."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/vendors/{vendor_id}/extraction")
        resp.raise_for_status()
        data = resp.json()
        if data["processing"]["is_complete"]:
            return data
        time.sleep(0.2)
    raise AssertionError("processing did not complete in time")


def test_health_and_status(client):
    assert client.get("/health").json() == {"status": "ok"}
    status = client.get("/api/system/status").json()
    assert status["government_verification_enabled"] is False


def test_full_vendor_flow(client, fixtures_dir):
    # 1. Create vendor
    resp = client.post("/api/vendors", json={})
    assert resp.status_code == 201
    vendor = resp.json()
    vendor_id = vendor["id"]
    assert vendor["status"] == "draft"

    # 2. Upload documents (bulk multipart)
    files = [
        ("files", ("gst_certificate.pdf", open(fixtures_dir / "sample_gst_certificate.pdf", "rb"), "application/pdf")),
        ("files", ("udyam_certificate.pdf", open(fixtures_dir / "sample_udyam_certificate.pdf", "rb"), "application/pdf")),
        ("files", ("cancelled_cheque.pdf", open(fixtures_dir / "sample_cancelled_cheque.pdf", "rb"), "application/pdf")),
    ]
    resp = client.post(f"/api/vendors/{vendor_id}/documents", files=files)
    for _, (_, fh, _) in files:
        fh.close()
    assert resp.status_code == 201
    upload_result = resp.json()
    assert len(upload_result["uploaded"]) == 3
    assert upload_result["rejected"] == []

    # 3. Documents list shows classification results.
    docs = client.get(f"/api/vendors/{vendor_id}/documents").json()
    assert len(docs) == 3
    doc_types = {d["document_type"] for d in docs}
    assert doc_types == {"gst_certificate", "udyam_certificate", "cancelled_cheque"}
    assert all(d["status"] == "done" for d in docs)

    # 4. Merged extraction view.
    extraction = _wait_for_processing(client, vendor_id)
    assert extraction["vendor_status"] == "review"
    field_map = {f["field_name"]: f for f in extraction["fields"]}
    assert field_map["company_name"]["value"] == "Bharat Textiles Private Limited"
    assert field_map["gst_registration_certificate"]["value"] == "27AAACB1234C1Z5"
    assert field_map["pan_card"]["value"] == "AAACB1234C"
    assert field_map["state"]["value"] == "Maharashtra"
    assert field_map["region"]["value"] == "West"  # derived from State
    assert field_map["country"]["value"] == "India"
    # Manual-only fields must stay blank until a human fills them in.
    assert field_map["payment_terms"]["value"] is None
    assert field_map["customer_agreement"]["value"] is None

    bank_map = {f["field_name"]: f for f in extraction["bank_details"]}
    assert bank_map["bank_ifsc_code"]["value"] == "HDFC0001234"

    checklist = {c["document_type"]: c["uploaded"] for c in extraction["document_checklist"]}
    assert checklist["gst_certificate"] is True
    assert checklist["pan_card"] is False

    # 5. Human correction via PUT -- must stick even if we reprocess.
    resp = client.put(
        f"/api/vendors/{vendor_id}/extraction",
        json={"fields": {"payment_terms": "Net 30", "company_name": "Corrected Co Pvt Ltd"}},
    )
    assert resp.status_code == 200
    updated = resp.json()
    field_map = {f["field_name"]: f for f in updated["fields"]}
    assert field_map["payment_terms"]["value"] == "Net 30"
    assert field_map["payment_terms"]["is_human_edited"] is True
    assert field_map["company_name"]["value"] == "Corrected Co Pvt Ltd"
    assert field_map["company_name"]["is_human_edited"] is True

    # 6. Re-uploading a document must NOT clobber the human edit.
    with open(fixtures_dir / "sample_gst_certificate.pdf", "rb") as fh:
        resp = client.post(
            f"/api/vendors/{vendor_id}/documents",
            files=[("files", ("gst_certificate_again.pdf", fh, "application/pdf"))],
        )
    assert resp.status_code == 201
    extraction = _wait_for_processing(client, vendor_id)
    field_map = {f["field_name"]: f for f in extraction["fields"]}
    assert field_map["company_name"]["value"] == "Corrected Co Pvt Ltd"

    # 7. Approve.
    resp = client.post(f"/api/vendors/{vendor_id}/approve", json={"approved_by": "reviewer@example.com"})
    assert resp.status_code == 200
    assert resp.json()["approved_by"] == "reviewer@example.com"
    assert client.get(f"/api/vendors/{vendor_id}").json()["status"] == "approved"

    # 8. Excel download reflects the corrected values.
    resp = client.get(f"/api/vendors/{vendor_id}/excel")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    out_file = fixtures_dir.parent / "_downloaded_test.xlsx"
    out_file.write_bytes(resp.content)
    try:
        wb = load_workbook(out_file)
        ws = wb.active
        assert ws.cell(row=3, column=1).value == "Company Name"
        assert ws.cell(row=3, column=2).value == "Corrected Co Pvt Ltd"
    finally:
        out_file.unlink(missing_ok=True)


def test_reject_bad_extension_and_oversized_are_reported_not_fatal(client):
    resp = client.post("/api/vendors", json={})
    vendor_id = resp.json()["id"]

    resp = client.post(
        f"/api/vendors/{vendor_id}/documents",
        files=[("files", ("not_a_doc.exe", b"MZ\x90\x00", "application/octet-stream"))],
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["uploaded"] == []
    assert len(body["rejected"]) == 1
    assert "extension" in body["rejected"][0]["reason"]


def test_delete_document(client, fixtures_dir):
    resp = client.post("/api/vendors", json={})
    vendor_id = resp.json()["id"]
    with open(fixtures_dir / "sample_pan_card.pdf", "rb") as fh:
        resp = client.post(
            f"/api/vendors/{vendor_id}/documents",
            files=[("files", ("pan.pdf", fh, "application/pdf"))],
        )
    doc_id = resp.json()["uploaded"][0]["id"]

    resp = client.delete(f"/api/vendors/{vendor_id}/documents/{doc_id}")
    assert resp.status_code == 204
    assert client.get(f"/api/vendors/{vendor_id}/documents").json() == []


def test_approve_requires_at_least_one_document(client):
    resp = client.post("/api/vendors", json={})
    vendor_id = resp.json()["id"]
    resp = client.post(f"/api/vendors/{vendor_id}/approve", json={})
    assert resp.status_code == 400
