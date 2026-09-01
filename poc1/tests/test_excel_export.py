"""Excel output: correct row order + values, per the exact layout spec."""
from openpyxl import load_workbook

from app.crud import create_vendor
from app.enums import EXCEL_FIELD_LABELS, EXCEL_FIELD_ORDER, ExtractionMethod, FieldName
from app.excel_export import export_vendor_excel
from app.extraction.merge import upsert_field
from app.extraction.common import FieldResult


def test_excel_row_order_and_values(db_session, tmp_path):
    vendor = create_vendor(db_session, company_name=None)

    upsert_field(db_session, vendor.id, FieldResult(FieldName.COMPANY_NAME, "Bharat Textiles Private Limited", ExtractionMethod.TEXT_LAYER, 0.9), None)
    upsert_field(db_session, vendor.id, FieldResult(FieldName.CITY, "Mumbai", ExtractionMethod.REGEX, 0.8), None)
    upsert_field(db_session, vendor.id, FieldResult(FieldName.STATE, "Maharashtra", ExtractionMethod.REGEX, 0.8), None)
    upsert_field(db_session, vendor.id, FieldResult(FieldName.ZIP_CODE, "400072", ExtractionMethod.REGEX, 0.9), None)
    upsert_field(db_session, vendor.id, FieldResult(FieldName.GST_REGISTRATION_CERTIFICATE, "27AAACB1234C1Z5", ExtractionMethod.REGEX, 0.97), None)
    db_session.commit()

    out_path = tmp_path / "vendor.xlsx"
    export_vendor_excel(db_session, vendor, out_path)

    wb = load_workbook(out_path)
    ws = wb.active

    # Row 3 is the first field row (row 1 = title, row 2 = spacer).
    rows = list(ws.iter_rows(min_row=3, max_row=2 + len(EXCEL_FIELD_ORDER), max_col=2, values_only=True))
    assert len(rows) == len(EXCEL_FIELD_ORDER)

    for (label, value), field_name in zip(rows, EXCEL_FIELD_ORDER):
        assert label == EXCEL_FIELD_LABELS[field_name]

    values_by_field = dict(zip(EXCEL_FIELD_ORDER, (v for _, v in rows)))
    assert values_by_field[FieldName.COMPANY_NAME] == "Bharat Textiles Private Limited"
    assert values_by_field[FieldName.CITY] == "Mumbai"
    assert values_by_field[FieldName.STATE] == "Maharashtra"
    assert values_by_field[FieldName.ZIP_CODE] == "400072"
    assert values_by_field[FieldName.GST_REGISTRATION_CERTIFICATE] == "27AAACB1234C1Z5"

    # Manual-only fields must stay blank -- nobody filled them in via PUT /extraction.
    # (openpyxl reads a cell that was written with "" back as None.)
    assert values_by_field[FieldName.CUSTOMER_AGREEMENT] in (None, "")
    assert values_by_field[FieldName.TYPE] in (None, "")
    assert values_by_field[FieldName.SALESPERSON] in (None, "")

    label_cell = ws.cell(row=3, column=1)
    assert label_cell.font.bold is True


def test_excel_respects_human_edited_manual_field(db_session, tmp_path):
    vendor = create_vendor(db_session, company_name=None)
    upsert_field(db_session, vendor.id, FieldResult(FieldName.TYPE, "Services", ExtractionMethod.MANUAL, 1.0), None)
    db_session.flush()
    from app.models import ExtractedField
    row = db_session.query(ExtractedField).filter_by(vendor_id=vendor.id, field_name=FieldName.TYPE.value).one()
    row.is_human_edited = True
    db_session.commit()

    out_path = tmp_path / "vendor2.xlsx"
    export_vendor_excel(db_session, vendor, out_path)
    wb = load_workbook(out_path)
    ws = wb.active
    idx = EXCEL_FIELD_ORDER.index(FieldName.TYPE)
    row_num = 3 + idx
    assert ws.cell(row=row_num, column=2).value == "Services"
